import openpyxl
from typing import Any, Optional
from datetime import datetime
from io import BytesIO


def validate_fields(headers: list[str], required: list[str]) -> list[str]:
    header_set = set(headers)
    missing = [f for f in required if f not in header_set]
    return missing


def parse_erp_row(row_data: dict[str, Any]) -> dict[str, Any]:
    order_time = None
    if row_data.get("时间"):
        try:
            if isinstance(row_data["时间"], datetime):
                order_time = row_data["时间"]
            elif isinstance(row_data["时间"], str):
                order_time = datetime.strptime(row_data["时间"], "%Y/%m/%d %H:%M:%S")
        except (ValueError, TypeError):
            pass

    balance_amount = None
    if row_data.get("结余"):
        raw = str(row_data["结余"]).replace("$", "").replace(",", "").strip()
        try:
            balance_amount = float(raw)
        except (ValueError, TypeError):
            balance_amount = None

    freight = _safe_float_optional(row_data.get("运费"))
    
    # 如果运费为0或者空，则服务费和打包费默认也为空
    if freight is None or freight == 0:
        service_fee = None
        packing_fee = None
    else:
        # 如果运费不为0或者空，则服务费默认为0，打包费默认为3元
        service_fee = 0.0
        packing_fee = 3.0
    
    total_cost = (freight or 0) + (service_fee or 0) + (packing_fee or 0)

    return {
        "erp_order_id": _safe_int(row_data.get("id")),
        "order_time": order_time,
        "source": str(row_data.get("来源") or ""),
        "currency": str(row_data.get("币种") or ""),
        "balance_amount": balance_amount,
        "asin": str(row_data.get("ASIN") or ""),
        "status": str(row_data.get("状态") or ""),
        "image_url": str(row_data.get("图片") or ""),
        "gross_weight": 0,
        "quantity": _safe_int(row_data.get("数量")),
        "channel": str(row_data.get("渠道") or "") if row_data.get("渠道") else "",
        "tracking_no": str(row_data.get("追踪号") or "") if row_data.get("追踪号") else "",
        "freight": freight,
        "service_fee": service_fee,
        "packing_fee": packing_fee,
        "total_cost": total_cost,
        "region": str(row_data.get("地区") or ""),
    }


def parse_logistics_row(row_data: dict[str, Any]) -> dict[str, Any]:
    return {
        "logistics_id": _safe_int(row_data.get("ID")),
        "channel_name": str(row_data.get("渠道名称") or ""),
        "country": str(row_data.get("国家") or ""),
        "weight": _safe_float(row_data.get("重量")),
        "salesman": str(row_data.get("业务员") or ""),
        "tracking_no": str(row_data.get("追踪号") or ""),
        "logistics_fee": _safe_float(row_data.get("物流费")),
        "service_fee": _safe_float(row_data.get("服务费")),
        "packing_fee": _safe_float(row_data.get("打包费")),
        "total_logistics_fee": _safe_float(row_data.get("物流总费用")),
        "status": str(row_data.get("状态") or ""),
        "logistics_status": str(row_data.get("物流状态") or ""),
        "time": _safe_datetime(row_data.get("时间")),
    }


def read_excel_to_dicts(file_content: bytes) -> tuple[list[str], list[dict[str, Any]]]:
    wb = openpyxl.load_workbook(BytesIO(file_content))
    ws = wb.active
    headers = [str(cell.value) if cell.value is not None else "" for cell in ws[1]]

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_dict = {}
        for i, header in enumerate(headers):
            if i < len(row):
                row_dict[header] = row[i]
            else:
                row_dict[header] = None
        rows.append(row_dict)

    return headers, rows


def _safe_float(val: Any) -> float:
    if val is None:
        return 0.0
    try:
        return float(val)
    except (ValueError, TypeError):
        return 0.0


def _safe_float_optional(val: Any) -> Optional[float]:
    if val is None:
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


def _safe_int(val: Any) -> int:
    if val is None:
        return 0
    try:
        return int(val)
    except (ValueError, TypeError):
        return 0


def _safe_datetime(val: Any) -> Optional[datetime]:
    if val is None:
        return None
    if isinstance(val, datetime):
        return val
    try:
        return datetime.strptime(str(val), "%Y-%m-%d %H:%M:%S")
    except (ValueError, TypeError):
        return None